from __future__ import annotations

import argparse
import csv
import math
import os
import statistics
import sys
from collections import defaultdict
from pathlib import Path

LOCAL_DEPS = Path(__file__).resolve().parent.parent / "python_deps"
if LOCAL_DEPS.exists():
    sys.path.insert(0, str(LOCAL_DEPS))
os.environ.setdefault("MPLCONFIGDIR", str(Path(__file__).resolve().parent / ".mplconfig"))

import matplotlib

matplotlib.use("Agg")
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.font_manager import FontProperties
from openpyxl import load_workbook

# Mandatory publication settings: sans-serif text and editable SVG glyphs.
plt.rcParams["font.family"] = "sans-serif"
plt.rcParams["font.sans-serif"] = ["Arial", "DejaVu Sans", "Liberation Sans"]
plt.rcParams["svg.fonttype"] = "none"
plt.rcParams["pdf.fonttype"] = 42
plt.rcParams["ps.fonttype"] = 42
plt.rcParams["font.size"] = 7
plt.rcParams["axes.linewidth"] = 0.7
plt.rcParams["figure.facecolor"] = "white"
plt.rcParams["axes.facecolor"] = "white"


GROUP_ORDER = ["HNO3", "H3PO4", "H2SO4", "HCl", "其他酸", "混合酸"]
GROUP_DISPLAY = {
    "HNO3": "HNO₃",
    "H3PO4": "H₃PO₄",
    "H2SO4": "H₂SO₄",
    "HCl": "HCl",
    "其他酸": "其他酸",
    "混合酸": "混合酸",
}
METRICS = [
    {
        "column": "比表面积（m²/g）",
        "label": "比表面积\n中位数",
        "unit": "(m² g⁻¹)",
        "format": lambda value: f"{value:.1f}" if value >= 100 else f"{value:.2f}",
    },
    {
        "column": "平均孔径（nm）",
        "label": "平均孔径\n中位数",
        "unit": "(nm)",
        "format": lambda value: f"{value:.2f}",
    },
    {
        "column": "总孔容（cm³/g）",
        "label": "总孔容\n中位数",
        "unit": "(cm³ g⁻¹)",
        "format": lambda value: f"{value:.3f}" if value < 0.1 else f"{value:.2f}",
    },
    {
        "column": "O（%）",
        "label": "氧含量\n中位数",
        "unit": "(%)",
        "format": lambda value: f"{value:.2f}",
    },
    {
        "column": "零电荷点（pH_pzc）",
        "label": "零电荷点\n中位数",
        "unit": "",
        "format": lambda value: f"{value:.2f}",
    },
]


def is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(float(value))


def load_group_medians(input_path: Path) -> tuple[np.ndarray, np.ndarray, dict[str, int]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook["Sheet1"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(rows)]
    index = {name: position for position, name in enumerate(headers)}

    required = ["酸组别", *[metric["column"] for metric in METRICS]]
    missing = [name for name in required if name not in index]
    if missing:
        raise KeyError(f"Sheet1 缺少必需字段: {', '.join(missing)}")

    values = defaultdict(lambda: defaultdict(list))
    group_counts = defaultdict(int)
    for row in rows:
        raw_group = row[index["酸组别"]]
        if raw_group is None:
            continue
        group = str(raw_group).strip()
        if group not in GROUP_ORDER:
            continue
        group_counts[group] += 1
        for metric in METRICS:
            value = row[index[metric["column"]]]
            if is_number(value):
                values[group][metric["column"]].append(float(value))

    medians = np.full((len(GROUP_ORDER), len(METRICS)), np.nan, dtype=float)
    valid_n = np.zeros_like(medians, dtype=int)
    for i, group in enumerate(GROUP_ORDER):
        for j, metric in enumerate(METRICS):
            observed = values[group][metric["column"]]
            valid_n[i, j] = len(observed)
            if observed:
                medians[i, j] = statistics.median(observed)
    return medians, valid_n, dict(group_counts)


def normalize_by_column(medians: np.ndarray) -> np.ndarray:
    normalized = np.full_like(medians, np.nan, dtype=float)
    for j in range(medians.shape[1]):
        column = medians[:, j]
        valid = np.isfinite(column)
        if not valid.any():
            continue
        lo = np.nanmin(column)
        hi = np.nanmax(column)
        if np.isclose(lo, hi):
            normalized[valid, j] = 0.5
        else:
            normalized[valid, j] = (column[valid] - lo) / (hi - lo)
    return normalized


def select_chinese_font() -> FontProperties:
    candidates = [
        Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
        Path("/System/Library/Fonts/STHeiti Light.ttc"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return FontProperties(fname=str(candidate))
    return FontProperties(family="sans-serif")


def write_source_csv(
    output_path: Path,
    medians: np.ndarray,
    normalized: np.ndarray,
    valid_n: np.ndarray,
    group_counts: dict[str, int],
) -> None:
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["acid_group", "group_sample_n", "metric", "valid_n", "median", "column_minmax_normalized"])
        for i, group in enumerate(GROUP_ORDER):
            for j, metric in enumerate(METRICS):
                writer.writerow([
                    group,
                    group_counts.get(group, 0),
                    metric["column"],
                    int(valid_n[i, j]),
                    "" if not np.isfinite(medians[i, j]) else f"{medians[i, j]:.10g}",
                    "" if not np.isfinite(normalized[i, j]) else f"{normalized[i, j]:.10g}",
                ])


def draw_figure(
    medians: np.ndarray,
    normalized: np.ndarray,
    group_counts: dict[str, int],
    output_stem: Path,
) -> None:
    font = select_chinese_font()
    cmap = LinearSegmentedColormap.from_list(
        "reference_blue",
        ["#F3F7FA", "#D9E9F2", "#A9CCE0", "#6EA6C9", "#2F6F9F", "#083B6F"],
    )
    cmap.set_bad("#F1F1EF")

    width_in = 118 / 25.4
    height_in = 150 / 25.4
    fig, ax = plt.subplots(figsize=(width_in, height_in), constrained_layout=False)
    masked = np.ma.masked_invalid(normalized)
    image = ax.imshow(masked, cmap=cmap, vmin=0, vmax=1, aspect="equal", interpolation="nearest")

    ax.set_xlim(-0.5, len(METRICS) - 0.5)
    ax.set_ylim(len(GROUP_ORDER) - 0.5, -0.5)
    ax.set_xticks(range(len(METRICS)))
    ax.set_yticks(range(len(GROUP_ORDER)))

    xlabels = [f"{metric['label']}\n{metric['unit']}" for metric in METRICS]
    ylabels = [f"{GROUP_DISPLAY[group]}\n(n = {group_counts.get(group, 0)})" for group in GROUP_ORDER]
    ax.set_xticklabels(xlabels, fontproperties=font, fontsize=7.2, linespacing=1.12)
    ax.set_yticklabels(ylabels, fontproperties=font, fontsize=7.2, linespacing=1.08)
    ax.xaxis.tick_top()
    ax.tick_params(axis="x", top=False, bottom=False, labeltop=True, labelbottom=False, length=0, pad=8)
    ax.tick_params(axis="y", length=0, pad=8)
    ax.set_ylabel("酸组别", fontproperties=font, fontsize=8, labelpad=18)

    ax.set_xticks(np.arange(-0.5, len(METRICS), 1), minor=True)
    ax.set_yticks(np.arange(-0.5, len(GROUP_ORDER), 1), minor=True)
    ax.grid(which="minor", color="white", linewidth=1.5)
    ax.tick_params(which="minor", top=False, bottom=False, left=False, right=False)
    for spine in ax.spines.values():
        spine.set_visible(False)

    for i in range(len(GROUP_ORDER)):
        for j, metric in enumerate(METRICS):
            value = medians[i, j]
            if not np.isfinite(value):
                ax.text(j, i, "—", ha="center", va="center", color="#777777", fontsize=7.2, fontproperties=font)
                continue
            shade = normalized[i, j]
            color = "white" if shade >= 0.60 else "#22313F"
            ax.text(
                j,
                i,
                metric["format"](value),
                ha="center",
                va="center",
                color=color,
                fontsize=7.5,
                fontweight="semibold",
                fontproperties=font,
            )

    cbar_ax = fig.add_axes([0.3075, 0.105, 0.625, 0.018])
    colorbar = fig.colorbar(image, cax=cbar_ax, orientation="horizontal")
    colorbar.set_ticks([0, 1])
    colorbar.set_ticklabels(["低", "高"])
    for label in colorbar.ax.get_xticklabels():
        label.set_fontproperties(font)
        label.set_fontsize(6.8)
    colorbar.ax.tick_params(length=0, pad=2)
    colorbar.outline.set_visible(False)
    fig.text(
        0.55,
        0.025,
        "颜色深浅表示该指标中位数大小。",
        ha="center",
        va="bottom",
        fontproperties=font,
        fontsize=6.2,
        color="#555555",
    )
    fig.subplots_adjust(left=0.27, right=0.97, top=0.78, bottom=0.19)

    fig.savefig(output_stem.with_suffix(".svg"), bbox_inches="tight", facecolor="white")
    fig.savefig(output_stem.with_suffix(".pdf"), bbox_inches="tight", facecolor="white")
    fig.savefig(output_stem.with_suffix(".png"), dpi=400, bbox_inches="tight", facecolor="white")
    fig.savefig(output_stem.with_suffix(".tiff"), dpi=600, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def write_contract(output_path: Path, group_counts: dict[str, int], valid_n: np.ndarray) -> None:
    total = sum(group_counts.values())
    minimum_cell_n = int(valid_n[valid_n > 0].min()) if np.any(valid_n > 0) else 0
    text = f"""# Figure contract and QA notes

- Core conclusion: The expanded dataset shows acid-group-specific median pore-structure, oxygen-content, and surface-charge profiles.
- Figure archetype: quantitative grid.
- Backend: Python (matplotlib) only.
- Final size: approximately 118 × 150 mm with square heatmap cells.
- Evidence: cell annotations are raw medians; color intensity is column-wise Min–Max normalization so variables with different units remain visually comparable.
- Statistics: descriptive medians only; no inferential test. Group labels report total sample count.
- Source data: Sheet1 of the provided updated workbook; {total} records across the six acid groups.
- Missingness: effective n differs by metric; the smallest non-zero cell n is {minimum_cell_n}. Exact cell-level n is retained in the source CSV.
- Reviewer risk: column-wise normalization supports within-metric comparison only; colors must not be compared quantitatively across different metric columns.
- Integrity: no image manipulation; the figure is generated directly from workbook values.
- Export QA: editable SVG, PDF, 400 dpi PNG, and 600 dpi TIFF are produced from the same Python figure.
"""
    output_path.write_text(text, encoding="utf-8")


def main() -> None:
    global METRICS
    parser = argparse.ArgumentParser(description="Draw the acid-group median heatmap from Sheet1.")
    parser.add_argument("--input", type=Path, required=True, help="Updated .xlsx workbook")
    parser.add_argument("--output-dir", type=Path, required=True, help="Directory for figure exports")
    parser.add_argument(
        "--legacy-aps-only",
        action="store_true",
        help="Also support the earlier four-response layout (without oxygen content).",
    )
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    if args.legacy_aps_only:
        METRICS = [METRICS[index] for index in (0, 1, 2, 4)]

    medians, valid_n, group_counts = load_group_medians(args.input)
    normalized = normalize_by_column(medians)
    output_name = "acid_group_median_heatmap_with_APS" if args.legacy_aps_only else "acid_group_median_heatmap_with_APS_and_O"
    source_name = "acid_group_median_heatmap_source_data.csv" if args.legacy_aps_only else "acid_group_median_heatmap_with_APS_and_O_source_data.csv"
    output_stem = args.output_dir / output_name
    write_source_csv(args.output_dir / source_name, medians, normalized, valid_n, group_counts)
    write_contract(args.output_dir / "figure_contract_and_QA.md", group_counts, valid_n)
    draw_figure(medians, normalized, group_counts, output_stem)

    print(f"rows_by_group={group_counts}")
    print(f"medians=\n{medians}")
    print(f"valid_n=\n{valid_n}")
    print(f"output_stem={output_stem}")


if __name__ == "__main__":
    main()
