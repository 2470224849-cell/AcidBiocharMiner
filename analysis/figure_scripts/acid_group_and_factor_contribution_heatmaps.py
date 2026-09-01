from __future__ import annotations

import argparse
import csv
import math
import os
import statistics
import sys
from collections import defaultdict
from pathlib import Path

LOCAL_DEPS = Path(__file__).resolve().parent.parent / "python_deps"
if LOCAL_DEPS.exists():
    sys.path.insert(0, str(LOCAL_DEPS))
os.environ.setdefault("MPLCONFIGDIR", str(Path(__file__).resolve().parent / ".mplconfig"))

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.font_manager import FontProperties
from openpyxl import load_workbook

plt.rcParams.update(
    {
        "font.family": "sans-serif",
        "font.sans-serif": ["Arial", "DejaVu Sans", "Liberation Sans"],
        "svg.fonttype": "none",
        "pdf.fonttype": 42,
        "ps.fonttype": 42,
        "font.size": 7,
        "axes.linewidth": 0.7,
        "figure.facecolor": "white",
        "axes.facecolor": "white",
    }
)

GROUP_ORDER = ["HNO3", "H3PO4", "H2SO4", "HCl", "其他酸", "混合酸"]
GROUP_DISPLAY = {
    "HNO3": "HNO₃",
    "H3PO4": "H₃PO₄",
    "H2SO4": "H₂SO₄",
    "HCl": "HCl",
    "其他酸": "其他酸",
    "混合酸": "混合酸",
}

TOP_METRICS = [
    {
        "column": "比表面积（m²/g）",
        "label": "比表面积\n中位数",
        "unit": "(m² g⁻¹)",
        "format": lambda value: f"{value:.1f}" if value >= 100 else f"{value:.2f}",
    },
    {
        "column": "平均孔径（nm）",
        "label": "平均孔径\n中位数",
        "unit": "(nm)",
        "format": lambda value: f"{value:.2f}",
    },
    {
        "column": "总孔容（cm³/g）",
        "label": "总孔容\n中位数",
        "unit": "(cm³ g⁻¹)",
        "format": lambda value: f"{value:.3f}" if value < 0.1 else f"{value:.2f}",
    },
    {
        "column": "O（%）",
        "label": "氧含量\n中位数",
        "unit": "(%)",
        "format": lambda value: f"{value:.2f}",
    },
    {
        "column": "零电荷点（pH_pzc）",
        "label": "零电荷点\n中位数",
        "unit": "",
        "format": lambda value: f"{value:.2f}",
    },
]

RESPONSES = [
    ("比表面积", "比表面积（m²/g）"),
    ("平均孔径", "平均孔径（nm）"),
    ("总孔容", "总孔容（cm³/g）"),
    ("氧含量", "O（%）"),
    ("零电荷点", "零电荷点（pH_pzc）"),
]

FACTORS = [
    ("酸类型", "acid_type"),
    ("标准化酸浓度", "acid_concentration"),
    ("酸处理时间", "acid_time"),
    ("酸处理温度", "acid_temperature"),
    ("改性顺序", "modification_order"),
    ("热解温度", "pyrolysis_temperature"),
    ("原料类别", "biomass_category"),
]

NUMERIC_FACTOR_COLUMNS = {
    "acid_concentration": "酸浓度（mol/L）",
    "acid_time": "酸处理时间（h）",
    "acid_temperature": "酸处理温度（°C）",
    "pyrolysis_temperature": "热解温度（°C）",
}

CATEGORICAL_FACTOR_COLUMNS = {
    "acid_type": "酸组别",
    "modification_order": "改性组别",
    "biomass_category": "生物质组别",
}


def select_chinese_font() -> FontProperties:
    candidates = [
        Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
        Path("/System/Library/Fonts/STHeiti Light.ttc"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return FontProperties(fname=str(candidate))
    return FontProperties(family="sans-serif")


def to_float(value: object) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def load_rows(input_path: Path) -> tuple[list[dict[str, object]], list[str]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook["Sheet1"]
    iterator = sheet.iter_rows(values_only=True)
    headers = ["" if value is None else str(value) for value in next(iterator)]
    required = {
        "酸组别",
        "改性组别",
        "生物质组别",
        *NUMERIC_FACTOR_COLUMNS.values(),
        *[column for _, column in RESPONSES],
    }
    missing = sorted(required.difference(headers))
    if missing:
        raise KeyError(f"Sheet1 缺少必需字段: {', '.join(missing)}")
    rows = []
    for values in iterator:
        if not any(value is not None for value in values):
            continue
        row = {header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header}
        rows.append(row)
    workbook.close()
    return rows, headers


def groupwise_standardize_concentration(rows: list[dict[str, object]]) -> None:
    grouped: dict[str, list[float]] = defaultdict(list)
    for row in rows:
        group = row.get("酸组别")
        concentration = to_float(row.get("酸浓度（mol/L）"))
        if group not in (None, "") and concentration is not None:
            grouped[str(group).strip()].append(concentration)

    stats: dict[str, tuple[float, float]] = {}
    for group, values in grouped.items():
        mean = statistics.fmean(values)
        sd = statistics.stdev(values) if len(values) > 1 else 0.0
        stats[group] = (mean, sd)

    for row in rows:
        group = row.get("酸组别")
        concentration = to_float(row.get("酸浓度（mol/L）"))
        standardized = None
        if group not in (None, "") and concentration is not None:
            mean, sd = stats.get(str(group).strip(), (math.nan, 0.0))
            if math.isfinite(mean) and sd > 0:
                standardized = (concentration - mean) / sd
        row["酸浓度_Cstar"] = standardized


def top_panel_data(rows: list[dict[str, object]]) -> tuple[np.ndarray, np.ndarray, dict[str, int]]:
    collected: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    counts: dict[str, int] = defaultdict(int)
    for row in rows:
        raw_group = row.get("酸组别")
        if raw_group in (None, ""):
            continue
        group = str(raw_group).strip()
        if group not in GROUP_ORDER:
            continue
        counts[group] += 1
        for metric in TOP_METRICS:
            value = to_float(row.get(metric["column"]))
            if value is not None:
                collected[group][metric["column"]].append(value)

    medians = np.full((len(GROUP_ORDER), len(TOP_METRICS)), np.nan)
    valid_n = np.zeros_like(medians, dtype=int)
    for i, group in enumerate(GROUP_ORDER):
        for j, metric in enumerate(TOP_METRICS):
            values = collected[group][metric["column"]]
            valid_n[i, j] = len(values)
            if values:
                medians[i, j] = statistics.median(values)
    return medians, valid_n, dict(counts)


def normalize_columns(values: np.ndarray) -> np.ndarray:
    normalized = np.full_like(values, np.nan, dtype=float)
    for j in range(values.shape[1]):
        column = values[:, j]
        valid = np.isfinite(column)
        if not valid.any():
            continue
        lo = float(np.nanmin(column))
        hi = float(np.nanmax(column))
        normalized[valid, j] = 0.5 if np.isclose(lo, hi) else (column[valid] - lo) / (hi - lo)
    return normalized


def zscore(values: np.ndarray) -> np.ndarray:
    mean = float(np.mean(values))
    sd = float(np.std(values, ddof=1)) if values.size > 1 else 0.0
    return np.zeros_like(values, dtype=float) if sd <= 0 else (values - mean) / sd


def one_hot(values: list[str]) -> np.ndarray:
    categories = sorted(set(values))
    if len(categories) <= 1:
        return np.empty((len(values), 0), dtype=float)
    return np.column_stack([[1.0 if value == category else 0.0 for value in values] for category in categories[1:]])


def fit_sse(blocks: dict[str, np.ndarray], y: np.ndarray, keep: list[str]) -> tuple[float, int]:
    matrices = [blocks[name] for name in keep if blocks[name].shape[1] > 0]
    x = np.column_stack([np.ones(y.size), *matrices]) if matrices else np.ones((y.size, 1))
    coefficients, _, rank, _ = np.linalg.lstsq(x, y, rcond=None)
    residuals = y - x @ coefficients
    return float(residuals @ residuals), int(rank)


def contribution_analysis(rows: list[dict[str, object]]) -> tuple[np.ndarray, list[dict[str, object]], dict[str, int]]:
    factor_keys = [key for _, key in FACTORS]
    contributions = np.full((len(FACTORS), len(RESPONSES)), np.nan)
    records: list[dict[str, object]] = []
    response_n: dict[str, int] = {}

    for response_index, (response_label, response_column) in enumerate(RESPONSES):
        complete: list[dict[str, object]] = []
        for row in rows:
            response = to_float(row.get(response_column))
            numeric_values = {
                "acid_concentration": to_float(row.get("酸浓度_Cstar")),
                "acid_time": to_float(row.get(NUMERIC_FACTOR_COLUMNS["acid_time"])),
                "acid_temperature": to_float(row.get(NUMERIC_FACTOR_COLUMNS["acid_temperature"])),
                "pyrolysis_temperature": to_float(row.get(NUMERIC_FACTOR_COLUMNS["pyrolysis_temperature"])),
            }
            categorical_values = {
                key: None if row.get(column) in (None, "") else str(row.get(column)).strip()
                for key, column in CATEGORICAL_FACTOR_COLUMNS.items()
            }
            if response is None or any(value is None for value in numeric_values.values()) or any(
                value is None for value in categorical_values.values()
            ):
                continue
            complete.append({"response": response, **numeric_values, **categorical_values})

        if len(complete) < 10:
            continue
        y = np.asarray([row["response"] for row in complete], dtype=float)
        response_n[response_label] = len(complete)
        blocks: dict[str, np.ndarray] = {}
        for key in NUMERIC_FACTOR_COLUMNS:
            values = np.asarray([row[key] for row in complete], dtype=float)
            blocks[key] = zscore(values).reshape(-1, 1)
        for key in CATEGORICAL_FACTOR_COLUMNS:
            values = [str(row[key]) for row in complete]
            blocks[key] = one_hot(values)

        sst = float(((y - np.mean(y)) ** 2).sum())
        full_sse, full_rank = fit_sse(blocks, y, factor_keys)
        full_r2 = max(0.0, 1.0 - full_sse / sst) if sst > 0 else 0.0
        deltas: dict[str, float] = {}
        partials: dict[str, float] = {}
        reduced_sse: dict[str, float] = {}
        for key in factor_keys:
            keep = [candidate for candidate in factor_keys if candidate != key]
            sse_without, _ = fit_sse(blocks, y, keep)
            reduced_sse[key] = sse_without
            delta = max(0.0, (sse_without - full_sse) / sst) if sst > 0 else 0.0
            partial = max(0.0, (sse_without - full_sse) / sse_without) if sse_without > 0 else 0.0
            deltas[key] = delta
            partials[key] = partial

        total_delta = sum(deltas.values())
        for factor_index, (factor_label, key) in enumerate(FACTORS):
            relative = 100.0 * deltas[key] / total_delta if total_delta > 0 else 0.0
            contributions[factor_index, response_index] = relative
            records.append(
                {
                    "response": response_label,
                    "response_column": response_column,
                    "complete_case_n": len(complete),
                    "full_model_rank": full_rank,
                    "full_model_r2": full_r2,
                    "factor": factor_label,
                    "factor_key": key,
                    "delta_r2": deltas[key],
                    "partial_r2": partials[key],
                    "relative_contribution_pct": relative,
                }
            )
    return contributions, records, response_n


def style_heatmap_axis(ax, rows: int, columns: int) -> None:
    ax.set_xlim(-0.5, columns - 0.5)
    ax.set_ylim(rows - 0.5, -0.5)
    ax.set_xticks(np.arange(-0.5, columns, 1), minor=True)
    ax.set_yticks(np.arange(-0.5, rows, 1), minor=True)
    ax.grid(which="minor", color="white", linewidth=1.5)
    ax.tick_params(which="minor", top=False, bottom=False, left=False, right=False)
    ax.xaxis.tick_top()
    ax.tick_params(axis="x", top=False, bottom=False, labeltop=True, labelbottom=False, length=0, pad=7)
    ax.tick_params(axis="y", length=0, pad=7)
    for spine in ax.spines.values():
        spine.set_visible(False)


def draw_combined_figure(
    output_stem: Path,
    medians: np.ndarray,
    median_normalized: np.ndarray,
    group_counts: dict[str, int],
    contributions: np.ndarray,
    response_n: dict[str, int],
) -> None:
    font = select_chinese_font()
    cmap = LinearSegmentedColormap.from_list(
        "reference_blue",
        ["#F3F7FA", "#D9E9F2", "#A9CCE0", "#6EA6C9", "#2F6F9F", "#083B6F"],
    )
    cmap.set_bad("#F1F1EF")

    fig = plt.figure(figsize=(118 / 25.4, 270 / 25.4))
    top_ax = fig.add_axes([0.27, 0.65, 0.70, 0.289])
    top_image = top_ax.imshow(
        np.ma.masked_invalid(median_normalized), cmap=cmap, vmin=0, vmax=1, aspect="equal", interpolation="nearest"
    )
    style_heatmap_axis(top_ax, len(GROUP_ORDER), len(TOP_METRICS))
    top_ax.set_xticks(range(len(TOP_METRICS)))
    top_ax.set_yticks(range(len(GROUP_ORDER)))
    top_ax.set_xticklabels(
        [f"{metric['label']}\n{metric['unit']}" for metric in TOP_METRICS],
        fontproperties=font,
        fontsize=7.0,
        linespacing=1.10,
    )
    top_ax.set_yticklabels(
        [f"{GROUP_DISPLAY[group]}\n(n = {group_counts.get(group, 0)})" for group in GROUP_ORDER],
        fontproperties=font,
        fontsize=7.0,
        linespacing=1.05,
    )
    top_ax.set_ylabel("酸组别", fontproperties=font, fontsize=8, labelpad=18)
    top_ax.tick_params(axis="x", which="both", top=False, bottom=False, labeltop=True, labelbottom=False, length=0)
    for i in range(len(GROUP_ORDER)):
        for j, metric in enumerate(TOP_METRICS):
            value = medians[i, j]
            if not np.isfinite(value):
                top_ax.text(j, i, "—", ha="center", va="center", color="#777777", fontsize=7, fontproperties=font)
                continue
            color = "white" if median_normalized[i, j] >= 0.60 else "#22313F"
            top_ax.text(
                j,
                i,
                metric["format"](value),
                ha="center",
                va="center",
                color=color,
                fontsize=7.2,
                fontweight="semibold",
                fontproperties=font,
            )
    fig.text(0.075, 0.955, "a", fontsize=9, fontweight="bold", ha="left", va="top")
    top_cbar_ax = fig.add_axes([0.3075, 0.612, 0.625, 0.010])
    top_cbar = fig.colorbar(top_image, cax=top_cbar_ax, orientation="horizontal")
    top_cbar.set_ticks([0, 1])
    top_cbar.set_ticklabels(["低", "高"])
    for label in top_cbar.ax.get_xticklabels():
        label.set_fontproperties(font)
        label.set_fontsize(6.5)
    top_cbar.ax.tick_params(length=0, pad=2)
    top_cbar.outline.set_visible(False)
    fig.text(
        0.62,
        0.590,
        "颜色深浅表示该指标中位数大小。",
        ha="center",
        va="center",
        fontproperties=font,
        fontsize=6.4,
        color="#555555",
    )

    bottom_ax = fig.add_axes([0.27, 0.17, 0.70, 0.337])
    contribution_vmax = max(10.0, math.ceil(float(np.nanmax(contributions)) / 5.0) * 5.0)
    bottom_image = bottom_ax.imshow(
        np.ma.masked_invalid(contributions),
        cmap=cmap,
        vmin=0,
        vmax=contribution_vmax,
        aspect="equal",
        interpolation="nearest",
    )
    style_heatmap_axis(bottom_ax, len(FACTORS), len(RESPONSES))
    bottom_ax.set_xticks(range(len(RESPONSES)))
    bottom_ax.set_yticks(range(len(FACTORS)))
    bottom_ax.set_xticklabels(
        [
            f"{r'pH$_{pzc}$' if label == 'pH_pzc' else label}\n(n = {response_n.get(label, 0)})"
            for label, _ in RESPONSES
        ],
        fontproperties=font,
        fontsize=7.0,
        linespacing=1.05,
    )
    bottom_ax.set_yticklabels([label for label, _ in FACTORS], fontproperties=font, fontsize=7.0)
    bottom_ax.set_ylabel("制备和改性因素", fontproperties=font, fontsize=8, labelpad=18)
    bottom_ax.tick_params(axis="x", which="both", top=False, bottom=False, labeltop=True, labelbottom=False, length=0)
    for i in range(len(FACTORS)):
        for j in range(len(RESPONSES)):
            value = contributions[i, j]
            if not np.isfinite(value):
                bottom_ax.text(j, i, "—", ha="center", va="center", color="#777777", fontsize=7, fontproperties=font)
                continue
            color = "white" if value / contribution_vmax >= 0.60 else "#22313F"
            display_value = "<0.1%" if 0 < value < 0.1 else f"{value:.1f}%"
            bottom_ax.text(
                j,
                i,
                display_value,
                ha="center",
                va="center",
                color=color,
                fontsize=7.1,
                fontweight="semibold",
                fontproperties=font,
            )
    fig.text(0.075, 0.535, "b", fontsize=9, fontweight="bold", ha="left", va="top")
    bottom_cbar_ax = fig.add_axes([0.3075, 0.115, 0.625, 0.010])
    bottom_cbar = fig.colorbar(bottom_image, cax=bottom_cbar_ax, orientation="horizontal")
    bottom_cbar.set_ticks([0, contribution_vmax / 2, contribution_vmax])
    bottom_cbar.set_ticklabels(["0", f"{contribution_vmax / 2:g}", f"{contribution_vmax:g}"])
    bottom_cbar.ax.tick_params(length=0, pad=2, labelsize=6.5)
    bottom_cbar.outline.set_visible(False)
    bottom_cbar.set_label("相对贡献（%）", fontproperties=font, fontsize=6.5, labelpad=2)
    fig.text(
        0.62,
        0.070,
        "格内数字为删除单因素后的独立 ΔR² 归一化占比；酸浓度已在同类酸内标准化。",
        ha="center",
        va="center",
        fontproperties=font,
        fontsize=6.1,
        color="#555555",
    )

    for suffix, dpi in [(".svg", None), (".pdf", None), (".png", 400), (".tiff", 600)]:
        fig.savefig(output_stem.with_suffix(suffix), dpi=dpi, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def write_contribution_csv(path: Path, records: list[dict[str, object]]) -> None:
    fields = [
        "response",
        "response_column",
        "complete_case_n",
        "full_model_rank",
        "full_model_r2",
        "factor",
        "factor_key",
        "delta_r2",
        "partial_r2",
        "relative_contribution_pct",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(records)


def write_protocol(path: Path, response_n: dict[str, int]) -> None:
    n_text = ", ".join(f"{key}={value}" for key, value in response_n.items())
    text = f"""# 多因素贡献分析方法

- 响应变量：SSA、APS、TPV、O 含量和 pH_pzc。
- 因素块：酸类型、酸浓度、酸处理时间、酸处理温度、改性顺序、热解温度和原料类别。
- 酸类型采用数据集中已归纳的 6 类“酸组别”；改性顺序采用已归纳的 6 类“改性组别”；原料类别采用“生物质组别”。
- 酸浓度先在每个酸组别内部按 C* = (C - 同类酸均值) / 同类酸样本标准差进行标准化。
- 每个响应单独采用完整案例；有效样本量：{n_text}。
- 模型为包含全部 7 个因素块的普通最小二乘多元模型；分类因素整体进行哑变量编码。
- 独立解释度定义为删除单个因素块后模型 R² 的下降量 ΔR²；热图格内数字为该 ΔR² 在 7 个因素 ΔR² 总和中的百分比。
- 该图用于描述数据集内的相对贡献，不代表因果效应；缺失值导致不同响应的建模样本量不同。
"""
    path.write_text(text, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Draw stacked acid-group median and factor-contribution heatmaps.")
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)

    rows, _ = load_rows(args.input)
    groupwise_standardize_concentration(rows)
    medians, valid_n, group_counts = top_panel_data(rows)
    median_normalized = normalize_columns(medians)
    contributions, records, response_n = contribution_analysis(rows)

    stem = args.output_dir / "acid_group_and_factor_contribution_heatmaps"
    draw_combined_figure(stem, medians, median_normalized, group_counts, contributions, response_n)
    write_contribution_csv(args.output_dir / "factor_relative_contribution_source_data.csv", records)
    write_protocol(args.output_dir / "factor_contribution_method.md", response_n)

    print(f"response_n={response_n}")
    print(f"contributions_pct=\n{contributions}")
    print(f"median_valid_n=\n{valid_n}")
    print(f"output_stem={stem}")


if __name__ == "__main__":
    main()
